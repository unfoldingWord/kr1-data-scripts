import os
import numpy as np
import requests
import pandas as pd
import mariadb
import json
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

load_dotenv()

AQUIFER_API_KEY = os.environ.get("AQUIFER_API_KEY")
AQUIFER_HEADERS = {"api-key": AQUIFER_API_KEY}
AQUIFER_BASE_URL = os.environ.get("AQUIFER_BASE_URL")

all_sli_categories = {
    "Bible Translation Aligned to Gk/Heb",
    "Exegetical Commentary",
    "Bible Translation Manual",
    "Gk/Heb Semantic Lexicons",
    "Gk/Heb Grammars",
    "Images, Maps, Videos",
    "Study Notes",
    "Translation Guide",
    "Comprehension Testing",
    "Bible Dictionary",
    "Bible Translation Source Text (audio preferred)",
    "Translation Glossary",
    "Foundational BT Training Videos",
    "Foundational Bible Stories"
}

# Define output directory and ensure it exists
OUTPUT_DIR = os.path.join(os.getcwd(), "output")
os.makedirs(OUTPUT_DIR, exist_ok=True)

EXCEL_FILE_PATH = os.path.join(OUTPUT_DIR, "kr1_report.xlsx")

db_config = {
    'host': os.environ.get("DB_HOST"),
    'user': os.environ.get("DB_USER"),
    'password': os.environ.get("DB_PASSWORD"),
    'database': os.environ.get("DATABASE")
}

GET_STRATEGIC_RESOURCES_QUERY = """
    SELECT 
        CONCAT(`language`, ' - [', bcp47, ']') AS `Strategic Language`, 
        bcp47 as language_code_2,
        iso_629_2 as language_code_3,
        resource_level AS `Resource Level`
    FROM sli_language_data sld
    WHERE resource_level IS NOT NULL AND resource_level > 0
    ORDER BY resource_level desc, `language` asc;
"""

GET_LANGUAGE_ENGAGEMENT_ISO_CODES = """
    select distinct sli.iso_629_2
    from language_engagements le 
    join uw_translation_products utp on utp.language_engagement_id = le.language_engagement_id
    join ietf_languages_codes ietf on ietf.ietf_id = le.ietf_id
    join sli_language_data sli on sli.iso_629_2 = ietf.iso_639_2
"""

dcs_aquifer_code_map = {
    "Translation Academy": "Bible Translation Manual",
    "Translation Words": "Translation Glossary",
    "TSV Translation Notes": "Translation Guide",
    "TSV Translation Questions": "Comprehension Testing",
    "Open Bible Stories": "Foundational Bible Stories",
    "Aligned Bible": "Bible Translation Aligned to Gk/Heb",
    "Hebrew Old Testament": "Bible Translation Source Text (audio preferred)",
    "Greek New Testament": "Bible Translation Source Text (audio preferred)",
}

INSERT_KR1_DATA = """
INSERT INTO kr1_progress_data (
    language_code,
    resource_name,
    resource_code,
    resource_owner,
    source,
    sli_category,
    resource_status
) VALUES (?, ?, ?, ?, ?, ?, ?)
"""


def fetch_aquifer_api_data(endpoint):
    """Fetch JSON data from Aquifer API"""
    url = f"{AQUIFER_BASE_URL}/{endpoint}"
    response = requests.get(url, headers=AQUIFER_HEADERS)
    if response.status_code != 200:
        raise Exception(f"API request to {endpoint} failed with status {response.status_code}")
    return response.json()


def get_aquifer_resources():
    """Fetch and process resource types"""
    aquifer_resource_data = fetch_aquifer_api_data("resources/types")
    df = pd.json_normalize(
        aquifer_resource_data,
        record_path=["collections"],
        meta=["type"],
        errors="ignore"
    ).rename(columns={
        "licenseInformation.title": "licenseInformation$title",
        "licenseInformation.copyright.dates": "licenseInformation$copyright$dates",
        "licenseInformation.copyright.holder.name": "licenseInformation$copyright$holder$name",
        "licenseInformation.copyright.holder.url": "licenseInformation$copyright$holder$url",
        "licenseInformation.licenses": "licenseInformation$licenses",
        "licenseInformation.showAdaptationNoticeForEnglish": "licenseInformation$showAdaptationNoticeForEnglish",
        "licenseInformation.showAdaptationNoticeForNonEnglish": "licenseInformation$showAdaptationNoticeForNonEnglish",
    })
    return df


def get_resource_collections(resource_codes):
    """Fetch collections for each resource and extract relevant language details."""
    collections = []

    for collection_code in resource_codes:
        collection_data = fetch_aquifer_api_data(f"resources/collections/{collection_code}")
        sli_category = collection_data.get("sliCategory", None)
        if sli_category:
            if sli_category == "Foundational Bible Stores":
                # hack because of misspelled word in aquifer
                sli_category = "Foundational Bible Stories"
            if sli_category not in all_sli_categories:
                raise ValueError(f'aquifer sli_category: {sli_category} not in hard coded category map.')
        else:
            raise AttributeError("missing sli category!!!")

        # Extract language details (normalize availableLanguages list)
        languages_data = pd.json_normalize(
            collection_data.get("availableLanguages", []),
            errors="ignore"
        )

        if not languages_data.empty:
            languages_data["code"] = collection_data.get("code")
            languages_data["display_name"] = collection_data.get("displayName")
            languages_data["resource_type"] = sli_category
            languages_data["source"] = "aquifer"
            languages_data["resource_owner"] = collection_data.get("licenseInfo", {}).get("copyright", {}).get("holder", {}).get("name")
        else:
            languages_data = pd.DataFrame([{
                "languageId": None,
                "languageCode": None,
                "displayName": None,
                "resourceItemCount": None,
                "code": collection_data.get("code"),
                "display_name": collection_data.get("displayName"),
                "resource_type": sli_category,
                "source": "aquifer",
                "resource_owner": collection_data.get("licenseInfo", {}).get("copyright", {}).get("holder", {}).get("name")
            }])

        collections.append(languages_data)

    combined_df = pd.concat(collections, ignore_index=True) if collections else pd.DataFrame()
    return combined_df


def get_languages():
    """Fetch language data from Aquifer"""
    return pd.DataFrame(fetch_aquifer_api_data("languages"))


def get_status(pct):
    if pct >= 90:
        return "Satisfied"
    elif pct > 0:
        return "In Progress"
    return ""


def get_bibles(aq_langs):
    """Fetch Bible resources and merge with languages"""
    aq_bibles = pd.DataFrame(fetch_aquifer_api_data("bibles"))
    aq_bibles["resource_status"] = "Satisfied"
    aq_bibles["source"] = "aquifer"
    aq_bibles["displayName"] = aq_bibles["name"]
    aq_bibles["resource_type"] = np.where(
        aq_bibles.get("hasGreekAlignment", False),
        "Bible Translation Aligned to Gk/Heb",
        "Bible Translation Source Text (audio preferred)"
    )
    aq_bibles["resource_owner"] = aq_bibles["licenseInfo"].apply(
        lambda x: x.get("copyright", {}).get("holder", {}).get("name", "") if isinstance(x, dict) else ""
    )
    # Merge with language data
    bibles = aq_bibles.merge(aq_langs, left_on="languageId", right_on="id", how="inner")
    return bibles[["englishDisplay", "name", "resource_status", "code", "source", "resource_type", "displayName", "resource_owner"]].rename(
        columns={"name": "resource_code", "code": "languageCode"}
    )


def generate_aquifer_resource_data():
    aq_resources = get_aquifer_resources()
    collections_df = get_resource_collections(aq_resources["code"])
    aq_langs = get_languages()
    bibles_df = get_bibles(aq_langs)

    # Fix: align languageId for merging
    aq_langs = aq_langs.rename(columns={"id": "languageId", "code": "languageCode"})
    collections_df = collections_df.drop(columns=["languageCode"], errors="ignore")  # drop existing one
    enriched = collections_df.merge(aq_langs[["languageId", "languageCode", "englishDisplay"]], on="languageId",
                                    how="left")

    # Get English baseline counts (languageId == 1)
    base_counts = (
        enriched[enriched["languageId"] == 1][["resource_code", "resourceItemCount"]]
        if "resource_code" in enriched.columns else
        enriched[enriched["languageId"] == 1][["code", "resourceItemCount"]].rename(columns={"code": "resource_code"})
    ).rename(columns={"resourceItemCount": "base_count"}).dropna()

    # Prepare working set
    enriched = enriched.rename(columns={"code": "resource_code"}).copy()
    enriched["resourceItemCount"] = enriched["resourceItemCount"].fillna(0).astype(int)

    # Merge in base count
    enriched = enriched.merge(base_counts, on="resource_code", how="left")

    # Compute completion_pct only where base_count is known
    enriched["completion_pct"] = enriched.apply(
        lambda row: round((row["resourceItemCount"] / row["base_count"]) * 100, 3)
        if pd.notnull(row["base_count"]) and row["base_count"] > 0 else None,
        axis=1
    )

    # Assign status
    def safe_get_status(row):
        if pd.isnull(row["base_count"]):
            return "Unknown"
        return get_status(row["completion_pct"])

    enriched["resource_status"] = enriched.apply(safe_get_status, axis=1)

    final_df = enriched[[
        "englishDisplay",
        "languageCode",
        "resource_code",
        "resource_status",
        "resource_type",
        "resource_owner",
        "displayName",
        "source"
    ]]

    # Append Bibles
    total = pd.concat([final_df, bibles_df], ignore_index=True)
    return total


def fetch_slr_data():
    connection = mariadb.connect(**db_config)
    try:
        cursor = connection.cursor(dictionary=True)
        cursor.execute(GET_STRATEGIC_RESOURCES_QUERY)
        results = cursor.fetchall()
        df = pd.DataFrame(results)
        return df
    finally:
        connection.close()


def get_language_engagement_iso_codes():
    connection = mariadb.connect(**db_config)
    try:
        cursor = connection.cursor(dictionary=True)
        cursor.execute(GET_LANGUAGE_ENGAGEMENT_ISO_CODES)
        results = cursor.fetchall()
        return set(row["iso_629_2"] for row in results if row["iso_629_2"])
    finally:
        connection.close()


def get_dcs_resource_status(data):
    prod_release_data = data.get("catalog", {}).get("prod", None)
    return "Satisfied" if prod_release_data else "In Progress"


def fetch_dcs_data():
    url = "https://git.door43.org/api/v1/repos/search?topic=tc-ready"
    response = requests.get(url)
    response.raise_for_status()
    data = response.json()
    rows = []
    if "data" in data:
        for obj in data["data"]:
            full_name = obj.get("full_name", "Unknown Name")
            subject = obj.get("subject", "N/A")
            abbreviation = obj.get("abbreviation", "N/A")
            url = obj.get("clone_url", "N/A")
            if subject in dcs_aquifer_code_map:
                sli_category = dcs_aquifer_code_map[subject]
            else:
                print(f'WARNING!!! repo {full_name} with a subject of {subject} and an abbreviation of {abbreviation} '
                      f'is not mapped and will not be included in delta report. Url is {url}')
                continue
            if sli_category not in all_sli_categories:
                raise ValueError(f'dcs sli_category: {sli_category} not in hard coded category map.')
            row = [
                f"{subject} ({full_name})",
                # hack below because dcs languages are Hebrew and Greek...
                "en" if subject == "Hebrew Old Testament" or subject == "Greek New Testament" else obj.get("language", "N/A"),
                "en" if subject == "Hebrew Old Testament" or subject == "Greek New Testament" else obj.get("language", "N/A").split("-", 1)[0],
                f"{subject} ({full_name})",
                sli_category,
                get_dcs_resource_status(obj),
                "unfoldingWord",
                "dcs"
            ]
            rows.append(row)
    df_columns = ["displayName", "englishDisplay", "languageCode", "resource_code", "resource_type", "resource_status", "resource_owner", "source"]
    df = pd.DataFrame(rows, columns=df_columns)
    return df


def calculate_status_from_resources(resources_df):
    # Priority order: Satisfied > In Progress > Unknown > ""
    priority = {"Satisfied": 3, "In Progress": 2, "Unknown": 1, "": 0, None: 0}
    highest = ("", 0)

    for _, row in resources_df.iterrows():
        status = row.get("resource_status")
        if priority.get(status, 0) > highest[1]:
            highest = (status, priority[status])

    return highest[0]


def save_to_fred(sl_resource_data, aquifer_dcs_data, headers):
    connection = mariadb.connect(**db_config)
    try:
        cursor = connection.cursor()
        # Clear existing data
        cursor.execute("DELETE FROM kr1_progress_data")

        for resource_type in headers:
            for _, row in sl_resource_data.iterrows():
                language_code_2 = row["language_code_2"]
                language_code_3 = row["language_code_3"]
                matching_resources = aquifer_dcs_data[
                    (aquifer_dcs_data["resource_type"] == resource_type) &
                    ((aquifer_dcs_data["languageCode"] == language_code_2) |
                     (aquifer_dcs_data["languageCode"] == language_code_3))
                ]
                language_code = language_code_3  # Use 3-letter ISO
                if matching_resources.empty:
                    cursor.execute(INSERT_KR1_DATA, (
                        language_code,
                        None,
                        None,
                        None,
                        None,
                        resource_type,
                        None
                    ))
                else:
                    for _, mr_row in matching_resources.iterrows():
                        resource_name = mr_row.get("displayName")
                        resource_code = mr_row.get("resource_code")
                        resource_owner = mr_row.get("resource_owner")
                        source = mr_row.get("source")
                        sli_category = mr_row.get("resource_type")
                        status = mr_row.get("resource_status")
                        cursor.execute(INSERT_KR1_DATA, (
                            language_code,
                            resource_name,
                            resource_code,
                            resource_owner,
                            source,
                            sli_category,
                            status
                        ))
        connection.commit()
    finally:
        connection.close()


def save_to_excel(sl_resource_data, aquifer_dcs_data, headers, file_path=EXCEL_FILE_PATH):
    """Export to Excel with full formatting: merged headers, resource shading, wrap text, etc."""

    strategic_iso_set = get_language_engagement_iso_codes()

    # Step 1: Expand headers
    expanded_headers = []
    for resource_type in headers:
        expanded_headers.append((resource_type, "Aquifer"))
        expanded_headers.append((resource_type, "DCS"))

    # Step 2: Build rows
    expanded_rows = []
    for _, row in sl_resource_data.iterrows():
        language_code_2 = row["language_code_2"]
        language_code_3 = row["language_code_3"]
        lang_name = row["Strategic Language"]
        if language_code_3 in strategic_iso_set:
            lang_name = f"{lang_name}**"
        new_row = [lang_name, language_code_2, language_code_3, row["Resource Level"]]

        for resource_type, source in expanded_headers:
            source_key = source.lower()
            matching_resources = aquifer_dcs_data[
                (aquifer_dcs_data["resource_type"] == resource_type) &
                (aquifer_dcs_data["source"] == source_key) &
                ((aquifer_dcs_data["languageCode"] == language_code_2) |
                 (aquifer_dcs_data["languageCode"] == language_code_3))
            ]

            final_status = calculate_status_from_resources(matching_resources)
            resource_lines = [f"---{res_row['resource_code']}" for _, res_row in matching_resources.iterrows()]
            cell_text = final_status + "\n" + "\n".join(sorted(resource_lines)) if resource_lines else final_status
            new_row.append(cell_text)

        expanded_rows.append(new_row)

    # Step 3: Create DataFrame
    core_columns = ["Strategic Language", "language_code_2", "language_code_3", "Resource Level"]
    all_columns = core_columns + expanded_headers
    expanded_df = pd.DataFrame(expanded_rows, columns=all_columns)
    expanded_df.drop(columns=["language_code_2", "language_code_3"], inplace=True)

    # Step 4: Write to Excel
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        expanded_df.to_excel(writer, sheet_name="Delta Report", index=False, header=False, startrow=2)
        worksheet = writer.sheets["Delta Report"]

        # Styles
        top_header_fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")
        aquifer_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        dcs_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        left_header_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
        wrap_alignment = Alignment(wrap_text=True, vertical="top", shrink_to_fit=False)
        bold_font = Font(bold=True)
        thin_border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )

        # Step 5: Headers
        already_merged = set()
        for col_idx, col in enumerate(expanded_df.columns, start=1):
            if isinstance(col, tuple):
                resource_type, source = col
                subheader_cell = worksheet.cell(row=2, column=col_idx, value=source)
                subheader_cell.font = bold_font
                subheader_cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="center")
                subheader_cell.border = thin_border
                subheader_cell.fill = aquifer_fill if source == "Aquifer" else dcs_fill

                if resource_type not in already_merged:
                    header_cell = worksheet.cell(row=1, column=col_idx, value=resource_type)
                    header_cell.font = bold_font
                    header_cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
                    header_cell.border = thin_border
                    header_cell.fill = top_header_fill
                    worksheet.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx + 1)
                    already_merged.add(resource_type)
            else:
                header_cell = worksheet.cell(row=1, column=col_idx, value=col)
                header_cell.font = bold_font
                header_cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
                header_cell.border = thin_border
                header_cell.fill = top_header_fill
                worksheet.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)

        # Step 6: Style data rows
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=3), start=3):
            for col_idx, cell in enumerate(row, start=1):
                cell.border = thin_border
                cell.alignment = wrap_alignment
                if col_idx == 1:
                    cell.fill = left_header_fill
                    cell.font = bold_font

        # Step 7: Auto column widths
        for col_cells in worksheet.columns:
            col_letter = get_column_letter(col_cells[0].column)
            col_idx = col_cells[0].column
            if col_idx == 1:
                worksheet.column_dimensions[col_letter].width = 35
            elif col_idx == 2:
                worksheet.column_dimensions[col_letter].width = 12
            else:
                worksheet.column_dimensions[col_letter].width = 70

        # Step 8: Footer rows
        from datetime import datetime
        last_data_row = len(expanded_df) + 3  # header is 2 rows; data starts on row 3
        footer_note_row = last_data_row + 1
        legend_row = last_data_row + 2

        for row_num, text in [(footer_note_row, f"Last Generated: {datetime.now().strftime('%m/%d/%Y %I:%M:%S %p')}"),
                              (legend_row, "** == uW engagements exist in this language.")]:
            worksheet.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
            for col in range(1, 3):
                footer_cell = worksheet.cell(row=row_num, column=col)
                footer_cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                footer_cell.border = thin_border
                footer_cell.alignment = Alignment(horizontal="left", vertical="center")
                if col == 1:
                    footer_cell.value = text

    print(f"Excel file saved at {file_path}")


slr_data = fetch_slr_data()
aquifer_data = generate_aquifer_resource_data()
dcs_data = fetch_dcs_data()
combined_data = pd.concat([aquifer_data, dcs_data], ignore_index=True)
sorted_slr_categories = all_sli_categories
save_to_excel(slr_data, combined_data, sorted(sorted_slr_categories), EXCEL_FILE_PATH)
save_to_fred(slr_data, combined_data, sorted(sorted_slr_categories))
print("KR1 delta report successfully generated and data saved to FRED!!!")
